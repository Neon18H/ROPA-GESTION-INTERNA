from datetime import timedelta
from decimal import Decimal

from django.db.models import DecimalField, ExpressionWrapper, F, Max, Q, Sum, Value
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView, View
from openpyxl import Workbook

from apps.common.mixins import RoleRequiredMixin
from apps.customers.models import Customer
from apps.finance.models import Expense
from apps.inventory.models import Stock
from apps.purchases.models import PurchaseItem, PurchaseOrder
from apps.sales.models import Sale
from apps.settings_app.models import StoreSettings

DECIMAL_12_2 = DecimalField(max_digits=12, decimal_places=2)
ZERO_DEC = Value(Decimal('0.00'), output_field=DECIMAL_12_2)


class ReportsView(RoleRequiredMixin, TemplateView):
    template_name = 'reports/index.html'
    allowed_roles = ('ADMIN', 'GERENTE')


class BaseXlsxReportView(RoleRequiredMixin, View):
    allowed_roles = ('ADMIN', 'GERENTE')
    filename = 'report.xlsx'

    def get_store_name(self, org):
        settings = StoreSettings.objects.using('settings_db').filter(organization_id=org.id).first()
        return settings.billing_legal_name or org.name if settings else org.name

    def build_workbook(self, org):
        raise NotImplementedError

    def get(self, request, *args, **kwargs):
        org = request.user.organization
        wb = self.build_workbook(org)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        wb.save(response)
        return response


class CustomerReportXlsxView(BaseXlsxReportView):
    filename = 'customers.xlsx'

    def build_workbook(self, org):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Clientes'
        ws.append([f'Tienda: {self.get_store_name(org)}'])
        ws.append([f'Generado: {timezone.localtime().strftime("%Y-%m-%d %H:%M") }'])
        ws.append([])
        ws.append(['Nombre', 'Teléfono', 'Email', 'Creado', 'Total compras', 'Última compra'])

        customers = Customer.objects.filter(organization=org).annotate(
            total_compras=Coalesce(
                Sum(
                    'sale__total',
                    filter=Q(sale__organization=org, sale__status=Sale.Status.PAID),
                    output_field=DECIMAL_12_2,
                ),
                ZERO_DEC,
                output_field=DECIMAL_12_2,
            ),
            ultima_compra=Max('sale__created_at', filter=Q(sale__organization=org)),
        ).order_by('name')

        for c in customers:
            ws.append([
                c.name,
                c.phone,
                c.email,
                c.created_at.strftime('%Y-%m-%d'),
                float(c.total_compras or 0),
                c.ultima_compra.strftime('%Y-%m-%d %H:%M') if c.ultima_compra else '',
            ])
        return wb


class InventoryReportXlsxView(BaseXlsxReportView):
    filename = 'inventory.xlsx'

    def build_workbook(self, org):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventario'
        ws.append([f'Tienda: {self.get_store_name(org)}'])
        ws.append([f'Generado: {timezone.localtime().strftime("%Y-%m-%d %H:%M") }'])
        ws.append([])
        ws.append(['SKU', 'Producto', 'Categoría', 'Marca', 'Talla', 'Color', 'Género', 'Barcode', 'Stock', 'Min alerta', 'Costo prom', 'Último costo', 'Valor inventario'])

        inv_value_expr = ExpressionWrapper(F('quantity') * Coalesce(F('avg_cost'), F('last_cost'), ZERO_DEC, output_field=DECIMAL_12_2), output_field=DECIMAL_12_2)
        stocks = Stock.objects.filter(variant__product__organization=org).select_related('variant__product__category', 'variant__product__brand').annotate(
            inventory_value=Coalesce(inv_value_expr, ZERO_DEC, output_field=DECIMAL_12_2)
        ).order_by('variant__product__sku')

        for s in stocks:
            v = s.variant
            p = v.product
            ws.append([
                p.sku,
                p.name,
                p.category.name if p.category else '',
                p.brand.name if p.brand else '',
                v.size,
                v.color,
                v.gender,
                v.barcode,
                s.quantity,
                s.min_alert,
                float(s.avg_cost or 0),
                float(s.last_cost or 0),
                float(s.inventory_value or 0),
            ])
        return wb


class FinanceReportXlsxView(BaseXlsxReportView):
    filename = 'finance.xlsx'

    def build_workbook(self, org):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Finanzas'
        today = timezone.localdate()
        start_30 = today - timedelta(days=29)
        month_start = today.replace(day=1)

        ws.append([f'Tienda: {self.get_store_name(org)}'])
        ws.append([f'Rango: {start_30} a {today}'])
        ws.append([])

        ws.append(['Ventas por día (30 días)'])
        ws.append(['Fecha', 'Total'])
        sales_daily = Sale.objects.filter(organization=org, status=Sale.Status.PAID, created_at__date__gte=start_30).annotate(day=TruncDate('created_at')).values('day').annotate(
            total=Coalesce(Sum('total', output_field=DECIMAL_12_2), ZERO_DEC, output_field=DECIMAL_12_2)
        ).order_by('day')
        for row in sales_daily:
            ws.append([row['day'].strftime('%Y-%m-%d'), float(row['total'])])

        ws.append([])
        ventas_mes = Sale.objects.filter(organization=org, status=Sale.Status.PAID, created_at__date__gte=month_start).aggregate(
            total=Coalesce(Sum('total', output_field=DECIMAL_12_2), ZERO_DEC, output_field=DECIMAL_12_2)
        )['total']
        gastos_mes = Expense.objects.filter(organization=org, date__gte=month_start).aggregate(
            total=Coalesce(Sum('amount', output_field=DECIMAL_12_2), ZERO_DEC, output_field=DECIMAL_12_2)
        )['total']
        ws.append(['Resumen mes actual'])
        ws.append(['Ventas mes', float(ventas_mes)])
        ws.append(['Gastos mes', float(gastos_mes)])

        ws.append([])
        ws.append(['Gastos'])
        ws.append(['Fecha', 'Categoría', 'Descripción', 'Monto'])
        for e in Expense.objects.filter(organization=org).order_by('-date')[:200]:
            ws.append([e.date.strftime('%Y-%m-%d'), e.category, e.description, float(e.amount or 0)])

        ws.append([])
        ws.append(['Compras confirmadas'])
        ws.append(['Proveedor', 'Fecha', 'Total'])
        purchase_expr = ExpressionWrapper(F('qty') * F('unit_cost'), output_field=DECIMAL_12_2)
        purchases = PurchaseItem.objects.filter(
            purchase__organization=org,
            purchase__status=PurchaseOrder.Status.RECEIVED,
        ).values('purchase_id', 'purchase__supplier__name', 'purchase__created_at').annotate(
            total=Coalesce(Sum(purchase_expr, output_field=DECIMAL_12_2), ZERO_DEC, output_field=DECIMAL_12_2)
        ).order_by('-purchase__created_at')
        for p in purchases:
            ws.append([p['purchase__supplier__name'], p['purchase__created_at'].strftime('%Y-%m-%d'), float(p['total'])])

        return wb
